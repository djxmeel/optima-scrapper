import io
import json
import os
import base64
import time
from copy import copy
from pathlib import Path

from PIL import Image
from io import BytesIO
import pandas as pd

import pypdf
import odoorpc
from pypdf.errors import PdfReadError
from odoorpc.error import RPCError
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
from reportlab.pdfgen import canvas
from selenium import webdriver

from utils.data_merger import DataMerger
from utils.odoo_import import OdooImport
from utils.util import Util


def login_odoo():
    #odoo_host = 'trialdb-final2.odoo.com'
    odoo_host = 'optimaluz.soluntec.net'
    odoo_protocol = 'jsonrpc+ssl'
    odoo_port = '443'

    #odoo_db = 'trialdb-final2'
    #odoo_login = 'itprotrial@outlook.com'
    #odoo_pass = 'itprotrial'

    odoo_db = 'Pruebas'
    odoo_login = 'productos@optimaluz.com'
    odoo_pass = '96c04503fc98aa4ffd90a9cf72ceb2d90d709b01'

    odoo = odoorpc.ODOO(odoo_host, protocol=odoo_protocol, port=odoo_port)
    # Authenticate with your credentials
    odoo.login(odoo_db, odoo_login, odoo_pass)

    return odoo

def change_internal_ref_odoo():
    odoo = login_odoo()

    product_model = odoo.env['product.template']

    # Fetch records in batches to avoid RPCerror
    batch_size = 200
    offset = 0
    products = []

    while True:
        product_ids = product_model.search([], offset=offset, limit=batch_size)
        if not product_ids:  # Exit the loop when no more records are found
            break

        products.extend(product_model.browse(product_ids))
        print(len(products))

        offset += batch_size

    for product in products:
        product_model.write(product.id, {'default_code': product.x_sku})


def ecommerce_filter_visibility_modifier(is_visible):
    odoo = login_odoo()
    product_attributes_model = odoo.env['product.attribute']

    attributes_ids = product_attributes_model.search([])

    product_attributes_model.write(attributes_ids, {'visibility': is_visible})


def rename_key_in_json_file(file_path, old_key, new_key):
    """
    Reads the JSON file, renames a specified key, and writes back the modified content.
    """
    with open(file_path, 'r') as f:
        data = json.load(f)

    # Check if the old key exists and rename it
    for product in data:
        if old_key in product:
            product[new_key] = product.pop(old_key)

    with open(file_path, 'w') as f:
        json.dump(data, f)


def rename_key_in_directory_jsons(directory, old_key, new_key):
    """
    Processes all JSON files in the specified directory and renames the old_key to new_key.
    """
    for filename in os.listdir(directory):
        if filename.endswith('.json'):
            file_path = os.path.join(directory, filename)
            rename_key_in_json_file(file_path, old_key, new_key)
            print(f"Processed {file_path}")


def process_sku_to_ref(directory):
    for filename in os.listdir(directory):
        if filename.endswith('.json'):
            file_path = os.path.join(directory, filename)

            with open(file_path, 'r') as f:
                print(f"OPENING {file_path}")
                data = json.load(f)

            # Check if the old key exists and rename it
            for product in data:
                if 'Sku' in product:
                    try:
                        if not product["Sku"].__contains__("VS"):
                            new_sku = int(product["Sku"])
                        else:
                            new_sku = int(product["Sku"][2:])
                    except ValueError:
                        product['default_code'] = f'VS{product["Sku"]}'
                        continue

                    product['default_code'] = f'VS{new_sku * 2}'
                    product['Sku'] = str(new_sku)
            with open(file_path, 'w') as f:
                json.dump(data, f)

            print(f"Processed {file_path}")


def process_ref_to_sku(directory):
    for filename in os.listdir(directory):
        if filename.endswith('.json'):
            file_path = os.path.join(directory, filename)

            with open(file_path, 'r') as f:
                print(f"OPENING {file_path}")
                data = json.load(f)

            # Check if the old key exists and rename it
            for product in data:
                if 'Sku' in product:
                    product['default_code'] = product['Sku']
                    del product['Sku']
                else:
                    product['default_code'] = str(int(int(product["default_code"][2:]) / 2))
            with open(file_path, 'w') as f:
                json.dump(data, f)

            print(f"Processed {file_path}")


def process_names_to_ref__clean_bad_skus(directory):
    for filename in os.listdir(directory):
        if filename.endswith('.json'):
            file_path = os.path.join(directory, filename)

            bad_sku_products = []

            with open(file_path, 'r') as f:
                print(f"OPENING {file_path}")
                data = json.load(f)

            # Check if the old key exists and rename it
            for product in data:
                if 'name' in product:
                    try:
                        product['name'] = f'[VS{int(product["default_code"]) * 2}] {product["name"].split("] ")[1]}'
                    except ValueError:
                        bad_sku_products.append(product)
                        print("REMOVING PRODUCT W/ BAD SKU: " + product['default_code'])
                        continue

            for product in bad_sku_products:
                data.remove(product)

            with open(file_path, 'w') as f:
                json.dump(data, f)

            print(f"Processed {file_path}")


def process_sku_to_ref_acc(directory):
    for filename in os.listdir(directory):
        if filename.endswith('.json'):
            file_path = os.path.join(directory, filename)

            with open(file_path, 'r') as f:
                print(f"OPENING {file_path}")
                data = json.load(f)

            # Check if the old key exists and rename it
            for product in data:
                if 'accesorios' in product:
                    for acc in product['accesorios']:
                        new_sku = int(acc["sku"][2:])

                        acc['default_code'] = f'VS{new_sku * 2}'
                        del acc['sku']

            with open(file_path, 'w') as f:
                json.dump(data, f)

            print(f"Processed {file_path}")


def process_ref_to_sku_acc(directory):
    for filename in os.listdir(directory):
        if filename.endswith('.json'):
            file_path = os.path.join(directory, filename)

            with open(file_path, 'r') as f:
                print(f"OPENING {file_path}")
                data = json.load(f)

            # Check if the old key exists and rename it
            for product in data:
                if 'accesorios' in product:
                    for acc in product['accesorios']:
                        if acc["default_code"].__contains__("VS"):
                            acc['default_code'] = str(int(int(acc["default_code"][2:]) / 2))
            with open(file_path, 'w') as f:
                json.dump(data, f)

            print(f"Processed {file_path}")


def field_update():
    odoo = login_odoo()

    product_model = odoo.env['product.template']
    product_attributes_model = odoo.env['product.attribute']

    product_ids = product_model.search([])

    for i, product_id in enumerate(product_ids):
        product_obj = product_model.browse(product_id)
        sku = product_obj.x_sku
        name = product_obj.name

        if sku:
            if sku[:2] == 'VS':
                sku = sku[2:]

            ref = f'VS{int(sku) * 2}'

            index = str(name).index(']')
            name = str(name).replace(name[:index + 1], f'[{ref}]')

            product_model.write(product_id, {
                'x_sku': sku,
                'default_code': ref,
                'name': name
            })

            print(f"{i}/{len(product_ids)} NEW NAME: {name}")

    print(f"Updated {len(product_ids)} products.")

def get_distinct_categs():
    categs = list(Util.load_json('data/vtac_italia/LINKS/PRODUCT_LINKS_CATEGORIES.json').values())
    categs += list(Util.load_json('data/vtac_uk/LINKS/PRODUCT_LINKS_CATEGORIES.json').values())

    distinct_categs = set()

    for categ_list in categs:
        for categ in categ_list:
            distinct_categs.add(categ)

    empty_translation_dict = {}

    for categ in sorted(list(distinct_categs)):
        empty_translation_dict[categ] = ''

    return empty_translation_dict


def stack_json_files_to_one(directory, output_file_path):
    """Generates a JSON file with all the products info from the specified directory."""
    products_info = []

    for file_path in Util.get_all_files_in_directory(directory):
        if file_path.endswith('.json'):
            with open(file_path, 'r') as f:
                print(f"OPENING {file_path}")
                data = json.load(f)

            products_info.extend(data)

    Util.dump_to_json(products_info, output_file_path)


def convert_xlsx_to_json(excel_file_path, json_file_path):
    # Load the XLSX file into a pandas DataFrame
    df = pd.read_excel(excel_file_path)

    # Convert the DataFrame to JSON
    json_data = df.to_json(orient='records', date_format='iso')

    # If you want to write the JSON data to a file, you could do so like this:
    with open(json_file_path, 'w') as json_file:
        json_file.write(json_data)

    return json_data


def upper_allproduct_names():
    products = OdooImport.browse_all_products_in_batches()
    odoo = login_odoo()
    product_model = odoo.env['product.template']

    for product in products:
        name = product.name.upper()
        product_model.write(product.id, {'name': name})
        print(f"UPDATED OLD: {product.name}\n NEW: {name}\n")


def delete_attachments(field, condition, value):
    odoo = login_odoo()
    product_model = odoo.env['product.template']
    attachments_model = odoo.env['ir.attachment']
    products_ids = product_model.search([(field, condition, value)])

    for product_id in products_ids:
        attachments = attachments_model.search([('res_id', '=', product_id)])
        attachments_model.unlink(attachments)
        print(f"DELETED ATTACHMENTS FOR PRODUCT: {product_id}")

def decode_and_save_b64_image(b64_string, output_folder, image_name):
    """Decode base64 string to image and save it"""
    image_data = base64.b64decode(b64_string)
    image = Image.open(BytesIO(image_data))
    image.save(os.path.join(output_folder, image_name))


def get_distinct_b64_imgs_from_json(dir_path, output_folder, field):
    """Process a JSON file with base64 encoded images and save DISTINCT images to a folder"""
    # Ensure output folder exists
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    files_paths = Util.get_all_files_in_directory(dir_path)
    b64_strings = []

    for file_path in files_paths:
        # Load the JSON data
        with open(file_path, 'r') as f:
            products_media = json.load(f)

        for product in products_media:
            if field not in product:
                continue

            # Get distinct base64 strings
            b64_strings.extend(product[field])
            print(len(b64_strings))

    distinct_b64_strings = list(set(b64_strings))
    print(len(distinct_b64_strings))

    # Decode and save each distinct image
    for index, b64_string in enumerate(distinct_b64_strings, 1):
        image_name = f'image_{index}.png'
        decode_and_save_b64_image(b64_string, output_folder, image_name)


def assign_public_categories(public_categories_path):
    odoo = login_odoo()
    public_categories_rows = Util.load_excel_columns_in_dictionary_list(public_categories_path)
    public_categories_model = odoo.env['product.public.category']
    product_model = odoo.env['product.template']

    for category_row in public_categories_rows:
        categ_ids = public_categories_model.search([('name', '=', str(category_row['CATEGORY ES']).strip())])
        product_id = product_model.search([('default_code', '=', str(category_row['SKU']).strip())])

        if categ_ids and product_id:
            print("ASSIGNING CATEGORY: " + category_row['CATEGORY ES'] + " TO SKU: " + str(category_row['SKU']))
            product_model.write(product_id[0], {'public_categ_ids': [(6, 0, categ_ids)]})
        else:
            if not categ_ids:
                print("CATEGORY NOT FOUND: " + category_row['CATEGORY ES'])
            if not product_id:
                print("SKU NOT FOUND: " + str(category_row['SKU']))


def delete_excel_rows(excel_file_path):
    """Delete rows from an Excel file."""
    data = Util.load_excel_columns_in_dictionary_list(excel_file_path)

    # load merged products and compare reference with rows with brand= V-TAC then delete the rows
    merged_products = [str(p['default_code']) for p in Util.load_data_in_dir('data/vtac_merged/PRODUCT_INFO')]

    for product in copy(data):
        if str(product['Brand']).strip() == 'V-TAC':
            if str(product['Referencia interna']) in merged_products:
                print("REMOVING (V-TAC AND EXISTS IN ODOO 16): " + str(product['Referencia interna'] + " " + str(product['Brand'])))
                data.remove(product)

    df = pd.DataFrame(data)
    df.to_excel('data/common/excel/NOT_ON_ODOO_16.xlsx', index=False)


def assign_public_categs_from_name():
    products = OdooImport.browse_all_products_in_batches('public_categ_ids', '=', False)

    for product in products:
        categs_names = Util.get_public_category_from_name(product.name, DataMerger.PUBLIC_CATEGORY_FROM_NAME_JSON_PATH)
        OdooImport.assign_public_categories(product.id, categs_names)


def delete_skus_in_odoo(skus_json_path):
    skus = Util.load_json(skus_json_path)
    odoo = login_odoo()
    product_model = odoo.env['product.template']

    for sku in skus["skus"]:
        product_id = product_model.search([('default_code', '=', sku)])
        if product_id:
            product_model.unlink(product_id)
            print(f"DELETED SKU: {sku}")


def archive_products_based_on_condition(attribute, condition, value):
    odoo = login_odoo()

    products_to_archive = odoo.env['product.template'].search([
        ('attribute_line_ids.attribute_id.name', '=', attribute),
        ('attribute_line_ids.value_ids.name', condition, value)
    ])

    for product in odoo.env['product.template'].browse(products_to_archive):
        product.write({'active': False})  # This archives the product in the database


def delete_all_unused_attributes_w_values():
    odoo = login_odoo()
    # Delete all product.attribute records
    attribute_ids = odoo.env['product.attribute'].search([])

    for attr_id in attribute_ids:
        attribute_value_ids = odoo.env['product.attribute.value'].search([('attribute_id', '=', attr_id)])

        for attr_value_id in attribute_value_ids:
            try:
                odoo.env['product.attribute.value'].unlink(attr_value_id)
                print(f'DELETED attribute.value {attr_value_id}')
            except RPCError:
                print(f'attribute.value {attr_value_id} is used.')
                continue

        try:
            odoo.env['product.attribute'].unlink(attr_id)
            print(f'DELETED attribute {attr_id}.')
        except RPCError:
            print(f'attribute {attr_id} is used.')
            continue


def set_all_prices(price, cost_also=False):
    products = OdooImport.browse_all_products_in_batches()
    odoo = login_odoo()
    product_model = odoo.env['product.template']

    for product in products:
        if cost_also:
            product_model.write(product.id, {'list_price': price, 'standard_price': price})
        else:
            product_model.write(product.id, {'list_price': price})
        print(f"UPDATED SKU: {product.default_code} PRICE: {price}")


def merge_excel_files(path1, path2, output_path, field, is_in=False, concat=True, additional_sku_filter_path=None):
    # Read the Excel files
    df1 = pd.read_excel(path1)
    df2 = pd.read_excel(path2)

    df1[field] = [str(d).replace('.0', '') for d in df1[field]]
    df2[field] = [str(d).replace('.0', '') for d in df2[field]]

    # Extract unique SKUs from the first DataFrame
    unique_skus = set(df1[field])

    # Filter the second DataFrame
    if is_in:
        df2_filtered = df2[df2[field].isin(unique_skus)]
    else:
        df2_filtered = df2[~df2[field].isin(unique_skus)]

    if additional_sku_filter_path and field == field:
        skus_to_skip = [sku for sku in Util.load_json(additional_sku_filter_path)["skus"]]
        df2_filtered = df2_filtered[~df2_filtered[field].isin(skus_to_skip)]

    if concat:
        # Combine the DataFrames
        result = pd.concat([df1, df2_filtered], ignore_index=True)
        result.to_excel(output_path, index=False)
    else:
        # Write the filtered DataFrame to an Excel file
        df2_filtered.to_excel(output_path, index=False)


# Used if you need only new products links but you don't have the last links file by comparing with Odoo skus
def new_links_only_odoo_comparator():
    driver = webdriver.Firefox()

    # IF ES JSON NOT GENERATED
    #extracted_links_es = Util.load_json('data/vtac_spain/LINKS/PRODUCTS_LINKS_ES.json')
    #extracted_links_skus_es = [{'url': link, 'sku': Util.get_sku_from_link_es(driver, link)} for link in extracted_links_es]
    #Util.dump_to_json(extracted_links_skus_es, 'data/vtac_spain/LINKS/extracted_links_skus_es.json')

    # IF ES JSON ALREADY GENERATED
    #extracted_links_skus_es = Util.load_json('data/vtac_spain/LINKS/extracted_links_skus_es.json')


    # IF UK JSON NOT GENERATED
    #extracted_links_uk = Util.load_json('data/vtac_uk/LINKS/PRODUCTS_LINKS_UK.json')
    #extracted_links_skus_uk = [{'url': link, 'sku': Util.get_sku_from_link_uk(driver, link)} for link in extracted_links_uk]
    #extracted_links_skus_uk = [entry for entry in extracted_links_skus_uk if entry['sku']] # remove empty entries
    #Util.dump_to_json(extracted_links_skus_uk, 'data/vtac_uk/LINKS/extracted_links_skus_uk.json')

    # IF UK JSON ALREADY GENERATED
    #extracted_links_skus_uk = Util.load_json('data/vtac_uk/LINKS/extracted_links_skus_uk.json')


    # IF ITA JSON NOT GENERATED
    extracted_links_ita = Util.load_json('data/vtac_italia/LINKS/PRODUCTS_LINKS_ITA.json')
    extracted_links_skus_ita = [{'url': link, 'sku': Util.get_sku_from_link_ita(driver, link)} for link in extracted_links_ita]
    Util.dump_to_json(extracted_links_skus_ita, 'data/vtac_italia/LINKS/extracted_links_skus_ita.json')

    # IF ITA JSON ALREADY GENERATED
    #extracted_links_skus_ita = Util.load_json('data/vtac_italia/LINKS/extracted_skus_ita.json')


    # IF skus_in_odoo16.json NOT GENERATED
    #skus_in_odoo16 = [prod.default_code for prod in OdooImport.browse_all_products_in_batches()]
    #Util.dump_to_json(skus_in_odoo16, 'data/common/json/skus_in_odoo16.json')

    # IF skus_in_odoo16.json ALREADY GENERATED
    skus_in_odoo16 = Util.load_json('data/common/json/skus_in_odoo16.json')


    #new_links_from_es = [entry['url'] for entry in extracted_links_skus_es if entry['sku'] not in skus_in_odoo16]
    #new_links_from_uk = [entry['url'] for entry in extracted_links_skus_uk if entry['sku'] not in skus_in_odoo16]
    new_links_from_ita = [entry['url'] for entry in extracted_links_skus_ita if entry['sku'] not in skus_in_odoo16]

    #Util.dump_to_json(new_links_from_es, 'data/vtac_spain/LINKS/NEW_PRODUCTS_LINKS_ES.json')
    #Util.dump_to_json(new_links_from_uk, 'data/vtac_uk/LINKS/NEW_PRODUCTS_LINKS_UK.json')
    Util.dump_to_json(new_links_from_ita, 'data/vtac_italia/LINKS/NEW_PRODUCTS_LINKS_ITA.json')

    driver.close()


def hardcode_field_odoo(field, value):
    odoo = login_odoo()
    products = OdooImport.browse_all_products_in_batches()
    product_model = odoo.env['product.template']

    for product in products:
        product_model.write(product.id, {field: value})
        print(f"UPDATED SKU: {product.default_code} {field}: {value}")


def find_duplicate_in_excel(excel_path, primary_key, output_path):
    # Read the Excel file
    df = pd.read_excel(excel_path)

    # Find duplicate SKUs. Keep='False' marks all duplicates as True
    duplicates = df[df.duplicated(primary_key, keep=False)]

    # Save the duplicates to a new Excel file
    duplicates.to_excel(output_path, index=False)

    return output_path


def get_price_variations_and_new_products_excel(primary_k, old_pricelist, new_pricelist, output_file):
    # Read the Excel files
    df1 = pd.read_excel(old_pricelist)
    df2 = pd.read_excel(new_pricelist)

    # Merge the dataframes on 'SKU' with a right join
    merged_df = pd.merge(df1, df2, on=primary_k, how='right', suffixes=('_file1', '_file2'), indicator=True)

    # Filter to keep rows with new SKUs and rows with different 'price' or 'promotions'
    filtered_df = merged_df[
        (merged_df['_merge'] == 'right_only') |
        ((merged_df['_merge'] == 'both') &
         ((merged_df['price_file1'] != merged_df['price_file2']) |
          (merged_df['promotions_file1'] != merged_df['promotions_file2'])))
        ]

    # Select columns from the second file only (excluding the merge indicator)
    relevant_columns = [col for col in merged_df.columns if '_file2' in col or col == primary_k or col == '_merge']
    filtered_df = filtered_df[relevant_columns]

    # Rename columns to remove suffix
    filtered_df.columns = filtered_df.columns.str.replace('_file2', '')

    # Write the result to a new Excel file
    filtered_df.to_excel(output_file, index=False)

    # Open the Excel file for formatting
    workbook = load_workbook(output_file)
    worksheet = workbook.active

    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    green_fill = PatternFill(start_color='05ff00', end_color='05ff00', fill_type='solid')

    # Apply conditional formatting
    for row in range(2, worksheet.max_row + 1):
        if worksheet[f'H{row}'].value == 'both':
            worksheet[f'H{row}'].fill = red_fill
            worksheet[f'H{row}'] = 'CAMBIO DE PRECIO O PROMOCIONES'
        elif worksheet[f'H{row}'].value == 'right_only':
            worksheet[f'H{row}'].fill = green_fill
            worksheet[f'H{row}'] = 'NUEVO PRODUCTO'

    # Save the workbook
    workbook.save(output_file)

def load_and_convert_images(input_directory, output_directory):
    # Create the output directory if it doesn't exist
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    # Track the unique base64 strings
    unique_images = set()

    # Iterate over all files in the input directory
    for filename in os.listdir(input_directory):
        if filename.endswith('.json'):
            filepath = os.path.join(input_directory, filename)
            with open(filepath, 'r') as file:
                data = json.load(file)
                for p in data:
                    # Extract base64 strings from 'icons' key
                    if 'icons' in p:
                        for image_b64 in p['icons']:
                            unique_images.add(image_b64)

    # Convert and save the unique images
    for i, image_b64 in enumerate(unique_images):
        image_data = base64.b64decode(image_b64)
        image = Image.open(BytesIO(image_data))
        image.save(os.path.join(output_directory, f'image_{i}.png'))


def create_products_from_excel(excel_path):
    products = Util.load_excel_columns_in_dictionary_list(excel_path)
    odoo = login_odoo()
    product_model = odoo.env['product.template']
    partner_model = odoo.env['res.partner']
    partner_id = partner_model.search([('name', '=', 'V-TAC Europe Ltd.')])[0]
    supplier_info_model = odoo.env['product.supplierinfo']
    categ_id = odoo.env['product.category'].search([('name', '=', 'Productos de iluminación')])[0]

    for product in products:
        product_id = product_model.create({
                        "default_code": str(product["SKU"]),
                        "name": product["Nombre"],
                        "standard_price": product["Compra"],
                        "list_price": 0,
                        "detailed_type": "product",
                        "invoice_policy": "delivery",
                        "categ_id": categ_id
                        #"product_brand_id": 1
                    })

        supplier_info_model.create({
            'partner_id': partner_id,
            'product_tmpl_id': product_id,
            'product_code': str(product["SKU"]),
            'price': product["Compra"],
            'min_qty': 1
        })

        print(f"CREATED SKU: {product['SKU']}")


def rename_files_in_subfolders(base_folder, new_name_pattern):
    """
    Renames all files in subfolders of the specified base folder.
    The new file names will follow the specified new_name_pattern.

    :param base_folder: Path to the base folder containing subfolders.
    :param new_name_pattern: A pattern for new file names, including an '{}' to be replaced with the original file name.
    """
    for root, dirs, files in os.walk(base_folder):
        for file in files:
            old_file_path = os.path.join(root, file)
            new_file_name = new_name_pattern.format(file)

            new_file_path = os.path.join(root, new_file_name)
            os.rename(old_file_path, new_file_path)
            print(f"Renamed '{old_file_path}' to '{new_file_path}'")

def remove_hyperlinks_from_pdf(input_pdf_path, output_pdf_path):
    # Read the input PDF
    try:
        input_pdf = pypdf.PdfReader(open(input_pdf_path, "rb"))
    except PdfReadError:
        print(f"ERROR READING {input_pdf_path}")
        return

    # Create a new PDF to write the modified content
    output_pdf = pypdf.PdfWriter()

    # Iterate through each page and remove annotations
    for i in range(len(input_pdf.pages)):
        page = input_pdf.pages[i]
        page[pypdf.generic.NameObject("/Annots")] = pypdf.generic.ArrayObject()
        output_pdf.add_page(page)

    # Write the modified content to a new file
    with open(output_pdf_path, "wb") as f:
        output_pdf.write(f)


def remove_hyperlinks_and_qr_code_from_pdfs(parent_folder, position, size):
    unedited_specsheets = []

    for root, dirs, files in os.walk(parent_folder):
        for file in files:
            if file.__contains__('FICHA_TECNICA_SKU'):
                continue
            input_pdf_path = os.path.join(root, file)
            output_pdf_path = os.path.join(root, f"FICHA_TECNICA_SKU_{file}")
            remove_hyperlinks_from_pdf(input_pdf_path, output_pdf_path)
            remove_elements_within_square(output_pdf_path, position, size, output_pdf_path)
            print(f"PROCESSED {input_pdf_path}")
            unedited_specsheets.append(input_pdf_path)

    for specsheet in unedited_specsheets:
        os.remove(specsheet)
        print(f"REMOVED {specsheet}")

    # Remove empty folders
    for root, dirs, files in os.walk(parent_folder):
        for dir in dirs:
            dir_path = os.path.join(root, dir)
            if not os.listdir(str(dir_path)):
                os.rmdir(dir_path)
                print(f"REMOVED EMPTY FOLDER: {dir_path}")


def create_white_square_overlay(position, size=(100, 100)):
    # Create a PDF in memory
    packet = io.BytesIO()
    c = canvas.Canvas(packet)
    c.setFillColorRGB(0, 0.807843137254902, 0.48627450980392156)  # White color
    c.rect(position[0], position[1], size[0], size[1], stroke=0, fill=1)
    c.save()

    packet.seek(0)
    return pypdf.PdfReader(packet)


def remove_elements_within_square(pdf_path, position, size, output_pdf_path):
    # Read the input PDF
    try:
        input_pdf = pypdf.PdfReader(open(pdf_path, "rb"))
    except PdfReadError:
        print(f"ERROR READING {pdf_path}")
        return
    except FileNotFoundError:
        print(f"ERROR READING {pdf_path}")
        return

    # Create a new PDF to write the modified content
    output_pdf = pypdf.PdfWriter()

    # Create the white square overlay PDF
    overlay_pdf = create_white_square_overlay(position, size)

    # Iterate through each page and apply the white square overlay
    for i in range(len(input_pdf.pages)):
        page = input_pdf.pages[i]
        if i == 0:
            page.merge_page(overlay_pdf.pages[0])
        output_pdf.add_page(page)

    # Write the modified content to a new file
    with open(output_pdf_path, "wb") as f:
        output_pdf.write(f)


def replace_name_files_in_subfolders(parent_folder, old_str, new_str):
    """
    Renames files in all subfolders of the given parent folder.

    Args:
    parent_folder (str): Path to the parent folder containing subfolders.
    old_str (str): The substring to be replaced in file names.
    new_str (str): The substring to replace with in file names.

    Returns:
    None: Files are renamed in place.
    """
    for subdir, _, files in os.walk(parent_folder):
        for file in files:
            old_file_path = Path(subdir) / file
            new_file_name = file.replace(old_str, new_str)
            new_file_path = Path(subdir) / new_file_name
            os.rename(old_file_path, new_file_path)
            pass


def encode_images_to_json(folder_path, output_path):
    for filename in os.listdir(folder_path):
        if filename.endswith(".png"):
            # Constructing the full file path
            file_path = os.path.join(folder_path, filename)

            # Reading the image and encoding it in base64
            with open(file_path, "rb") as image_file:
                encoded_string = base64.b64encode(image_file.read()).decode('utf-8')

            # Creating a dictionary with the filename as key and encoded image as value
            data = {filename.split('.png')[0]: encoded_string}

            # Writing the JSON file
            json_filename = os.path.splitext(filename)[0] + '.json'
            with open(os.path.join(output_path, json_filename), 'w') as json_file:
                json.dump(data, json_file)

def find_duplicate_skus(file_path):
    # Load the Excel file
    df = pd.read_excel(file_path, engine='openpyxl')

    # Check if 'SKU' column exists
    if 'SKU' not in df.columns:
        print("The 'SKU' column does not exist in the Excel file.")
        return

    # Find duplicates in the 'SKU' column
    duplicates = df[df.duplicated('SKU', keep=False)]  # keep=False marks all duplicates as True

    if duplicates.empty:
        print("No duplicate SKUs found.")
    else:
        print("Duplicate SKUs found:")
        print(duplicates[['SKU']])


def match_and_write_to_excel(json_file_path1, json_file_path2, pricelist_path, output_excel_path):
    # Load JSON files
    with open(json_file_path1, 'r') as file:
        data1 = json.load(file)
    with open(json_file_path2, 'r') as file:
        data2 = json.load(file)

    pricelist_data = Util.load_excel_columns_in_dictionary_list(pricelist_path)

    # Assuming data1 and data2 are lists of dictionaries
    # Convert them to dictionaries with IDs as keys for easier access
    data1_dict = {str(item['sku']): item for item in data1}
    data2_dict = {str(item['sku']): item for item in data2}
    pricelist_dict = {str(item['SKU']): item for item in pricelist_data}

    # Find matching IDs and combine their data
    matched_data = []
    for id in data1_dict:
        coste = 'null'

        if id in pricelist_dict:
            coste = pricelist_dict[id]['PRECIO COMPRA']

        if id in data2_dict:
            # Assuming you want to combine all fields, update with data2's fields
            matched_data.append({
                'sku': id,
                'Stock Italia 1': data1_dict[id]['stock_ita'],
                'Stock Italia 2': data2_dict[id]['stock_ita'],
                'Stock buyled 1': data1_dict[id]['stock_buyled'],
                'Stock buyled 2': data2_dict[id]['stock_buyled'],
                'Precio buyled 1': data1_dict[id]['price'],
                'Precio buyled 2': data2_dict[id]['price'],
                'Precio coste': coste,
            })
        else:
            # Add the data from data1
            matched_data.append({
                'sku': id,
                'Stock Italia 1': data1_dict[id]['stock_ita'],
                'Stock Italia 2': 'null',
                'Stock buyled 1': data1_dict[id]['stock_buyled'],
                'Stock buyled 2': 'null',
                'Precio buyled 1': data1_dict[id]['price'],
                'Precio buyled 2': 'null',
                'Precio coste': coste,
            })

    # Convert matched data to a pandas DataFrame
    df = pd.DataFrame(matched_data)

    # Write the DataFrame to an Excel file
    df.to_excel(output_excel_path, index=False)


# Example usage
pricelist_path = 'data/common/excel/pricelist_compra_coste.xlsx'
json_file_path1 = 'data/buyled_stocks/buyled_stocks_3450.json'
json_file_path2 = 'data/buyled_stocks/buyled_stocks_3450 - Copy.json'
output_excel_path = 'data/buyled_stocks/output.xlsx'

#match_and_write_to_excel(json_file_path1, json_file_path2, pricelist_path, output_excel_path)

# Uncomment the line below to run the function with your file paths
# match_and_write_to_excel(json_file_path1, json_file_path2, output_excel_path)


# position = (490, 740)  # X, Y coordinates
# size = (80, 80)  # Width, Height of the square
# parent_folder = "data/vtac_uk/SPEC_SHEETS"
# remove_hyperlinks_and_qr_code_from_pdfs(parent_folder, position, size)

# Example usage
# file_path = 'data/common/excel/product_icons.xlsx'
# find_duplicate_skus(file_path)

# Usage
#encode_images_to_json('data/common/icons/icons_catalog_Q1_2024', 'data/common/icons/icons_b64')


#replace_name_files_in_subfolders('data/vtac_spain/PRODUCT_PDF', 'Technical Specifications', 'Especificaciones Técnicas')

# position = (490, 740)  # X, Y coordinates
# size = (80, 80)  # Width, Height of the square
# parent_folder = "data/vtac_uk/SPEC_SHEETS"
# remove_hyperlinks_and_qr_code_from_pdfs(parent_folder, position, size)

# Example usage
# dir_path1 = 'data/common/icons/icons_catalog_Q1_2024'
# dir_path2 = 'data/common/icons/distinct_icons_ita'
# match_images(dir_path1, dir_path2)


#create_products_from_excel("data/common/excel/AvideEntac_con_movimientos_OK.xlsx")

# Example usage
# load_and_convert_images('data/vtac_italia/PRODUCT_MEDIA', 'data/vtac_italia/distinct_icons')


#Example usage
# get_price_variations_and_new_products_excel(
#      'sku',
#      'data/common/excel/to_compare/old.xlsx',
#      'data/common/excel/to_compare/new.xlsx',
#      'data/common/excel/to_compare/output.xlsx')


# Example usage :
#find_duplicate_in_excel('data/common/excel/productos_odoo_15.xlsx', 'SKU', 'data/common/excel/duplicates.xlsx')
#find_duplicate_in_excel('C:/Users/Djamel/Downloads/Producto_product.product.xlsx', 'SKU', 'data/common/excel/duplicates.xlsx')

# merge_excel_files(
#     'data/common/excel/to_compare/pricelist.xlsx',
#     "data/common/excel/to_compare/en O15 y no en pricelist.xlsx",
#     'data/common/excel/to_compare/output.xlsx',
#     'SKU',
#     True,
#     False,
#     "data/common/json/SKUS_TO_SKIP.json"
# )

#delete_excel_rows("data/common/excel/productos_odoo_15.xlsx")

#get_distinct_b64_imgs_from_json('data/vtac_merged/PRODUCT_MEDIA', 'data/unique_icons', 'icons')
# Util.dump_to_json(get_distinct_categs(), Util.PUBLIC_CATEGORIES_TRANSLATION_PATH)

#ecommerce_filter_visibility_modifier('hidden')

#convert_xlsx_to_json('data/common/VTAC_ES_PUBLIC_CATEGORIES.xlsx', 'data/common/PUBLIC_CATEGORIES.json')


#process_ref_to_sku(DataMerger.MERGED_PRODUCT_INFO_DIR_PATH)
#process_ref_to_sku(DataMerger.MERGED_PRODUCT_MEDIA_DIR_PATH)

#process_ref_to_sku_acc(DataMerger.MERGED_PRODUCT_INFO_DIR_PATH)
#change_internal_ref_odoo()

#stack_json_files_to_one("data/buyled_stocks - copia", "data/buyled_stocks - copia/buyled_stocks_all.json")

#delete_attachments('x_url', 'ilike', 'italia')

#assign_public_categories('data/common/excel/public_category_sku_Q1_2024.xlsx')
#assign_public_categs_from_name()
#assign_public_categories('data/common/excel/public_category_manual.xlsx')

#delete_skus_in_odoo('data/common/json/SKUS_TO_SKIP.json')

#set_all_prices(0, True)

#upper_allproduct_names()

#hardcode_field_odoo('product_brand_id', 1)

#delete_all_unused_attributes_w_values()

#archive_products_based_on_condition('Tipo de casquillo', '=', 'B22')

#new_links_only_odoo_comparator()
